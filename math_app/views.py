import openpyxl
import os
import pandas as pd
from django.shortcuts import render
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from win32com import client
import pythoncom
import formulas
from django.http import HttpResponse
from django.views.decorators.clickjacking import xframe_options_deny
from django.views.decorators.clickjacking import xframe_options_sameorigin


def index(request):
    wb = openpyxl.load_workbook(os.getcwd() + r'\excel\book.xlsx')
    sheet1 = wb['Sheet1']
    num1 = request.GET.get('num1')
    num2 = request.GET.get('num2')

    # save data in excel
    sheet1.cell(column=1, row=2, value=num1)
    sheet1.cell(column=2, row=2, value=num2)
    wb.save(f'{os.getcwd()}/excel/book.xlsx')
    wb.close()

    # انشاء نسخة من ملف الاكسيل
    xl_model = formulas.ExcelModel().loads(f'{os.getcwd()}/excel/book.xlsx').finish()
    xl_model.calculate()
    # if PermissionError:
    #     with open('../~$BOOK.XLSX','+wb') as f:
    #         pass
    #     if f.closed:
    #         return HttpResponse('True')
    #     else:
    #         return HttpResponse('False')
    xl_model.write(dirpath='./')
    # القراءة من نسخة ملف الاكسيل
    # ملاحظة النسخة تكون بالاحرف الكبيرة
    wb = openpyxl.load_workbook(f'{os.getcwd()}/BOOK.XLSX', data_only=True)
    sheet1 = wb['SHEET1']
    # Reading from Excel and displaying it in the browser
    df = pd.read_excel(r'BOOK.XLSX', sheet_name='SHEET1', usecols='A:L')
    number1 = df.loc[0]['number 1']
    number2 = df.loc[0]['number 2']
    sol = df.loc[0]['sol']

    context = {'num1': number1,
               'num2': number2,
               'sol': sol}

    return render(request, 'html_pages/index.html', context)


def some_view(request):
    # إنشاء ذاكرة مؤقتة لتخزين البيانات قبل حفظهافي ملف
    buffer = io.BytesIO()

    # إنشاء كائن الملف باستخدام مكتبة
    # reportlab
    p = canvas.Canvas(buffer)

    p.drawString(100, 750, 'PDF file')

    # غلق كائن الملف وحفظه
    p.showPage()
    p.save()

    # إرسال الملف إلى العميل لتحميله
    buffer.seek(0)
    return FileResponse(open(f'{os.getcwd()}/templates/PDF/file.pdf', 'rb'), buffer, as_attachment=True,
                        filename='math.pdf')


@xframe_options_sameorigin
def pdf_view(request):
    """هذه الدالة افظل من الدالة pdf_file في عرض ملف pdf لانها لاتقم بفتح ملف excel"""
    with open(f'{os.getcwd()}/templates/PDF/file.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename=file.pdf'
        return response


@xframe_options_sameorigin
def pdf_file(request) -> object:
    excel = client.Dispatch("Excel.Application", pythoncom.CoInitialize())
    # Read Excel File
    sheets = excel.Workbooks.Open(f'{os.getcwd()}/BOOK.XLSX')
    work_sheets = sheets.Worksheets[0]
    # Convert into PDF File
    work_sheets.ExportAsFixedFormat(0, f'{os.getcwd()}/templates/PDF/file.pdf')
    # close Excel
    sheets.Close()
    return FileResponse(open(f'{os.getcwd()}/templates/PDF/file.pdf', 'rb'), content_type='application/pdf')


def excel1(request):
    wb = openpyxl.load_workbook(os.getcwd() + r'\excel\excel1.xlsx')
    sheet1 = wb['Sheet1']
    num1 = request.GET.get('num1')
    num2 = request.GET.get('num2')

    # save data in excel
    sheet1.cell(column=1, row=2, value=num1)
    sheet1.cell(column=2, row=2, value=num2)
    wb.save(f'{os.getcwd()}/excel/excel1.xlsx')
    wb.close()

    # انشاء نسخة من ملف الاكسيل
    xl_model = formulas.ExcelModel().loads(f'{os.getcwd()}/excel/excel1.xlsx').finish()
    xl_model.calculate()

    xl_model.write(dirpath='./')
    # القراءة من نسخة ملف الاكسيل
    # ملاحظة النسخة تكون بالاحرف الكبيرة
    wb = openpyxl.load_workbook(f'{os.getcwd()}/EXCEL1.XLSX', data_only=True)
    sheet1 = wb['SHEET1']
    # Reading from Excel and displaying it in the browser
    df = pd.read_excel(r'EXCEL1.XLSX', sheet_name='SHEET1', usecols='A:L')

    number1 = df.loc[0]['number 1']
    number2 = df.loc[0]['number 2']
    sol = df.loc[0]['sol']

    context = {'num1': number1,
               'num2': number2,
               'sol': sol}

    return render(request, 'html_pages/excel.html', context)


def excel2(request):
    wb = openpyxl.load_workbook(os.getcwd() + r'\excel\excel2.xlsx')
    sheet1 = wb['Sheet1']
    num1 = request.GET.get('num1')
    num2 = request.GET.get('num2')

    # save data in excel
    sheet1.cell(column=1, row=2, value=num1)
    sheet1.cell(column=2, row=2, value=num2)
    wb.save(f'{os.getcwd()}/excel/excel2.xlsx')
    wb.close()

    # انشاء نسخة من ملف الاكسيل
    xl_model = formulas.ExcelModel().loads(f'{os.getcwd()}/excel/excel2.xlsx').finish()
    xl_model.calculate()

    xl_model.write(dirpath='./')
    # القراءة من نسخة ملف الاكسيل
    # ملاحظة النسخة تكون بالاحرف الكبيرة
    wb = openpyxl.load_workbook(f'{os.getcwd()}/EXCEL2.XLSX', data_only=True)
    sheet1 = wb['SHEET1']
    # Reading from Excel and displaying it in the browser
    df = pd.read_excel(r'EXCEL2.XLSX', sheet_name='SHEET1', usecols='A:L')

    number1 = df.loc[0]['number 1']
    number2 = df.loc[0]['number 2']
    sol = df.loc[0]['sol']

    context = {'num1': number1,
               'num2': number2,
               'sol': sol}

    return render(request, 'html_pages/excel2.html', context)


def excel3(request):
    wb = openpyxl.load_workbook(os.getcwd() + r'\excel\excel3.xlsx')
    sheet1 = wb['Sheet1']
    num1 = request.GET.get('num1')
    num2 = request.GET.get('num2')

    # save data in excel
    sheet1.cell(column=1, row=2, value=num1)
    sheet1.cell(column=2, row=2, value=num2)
    wb.save(f'{os.getcwd()}/excel/excel3.xlsx')
    wb.close()

    # انشاء نسخة من ملف الاكسيل
    xl_model = formulas.ExcelModel().loads(f'{os.getcwd()}/excel/excel3.xlsx').finish()
    xl_model.calculate()

    xl_model.write(dirpath='./')
    # القراءة من نسخة ملف الاكسيل
    # ملاحظة النسخة تكون بالاحرف الكبيرة
    wb = openpyxl.load_workbook(f'{os.getcwd()}/EXCEL3.XLSX', data_only=True)
    sheet1 = wb['SHEET1']
    # Reading from Excel and displaying it in the browser
    df = pd.read_excel(r'EXCEL3.XLSX', sheet_name='SHEET1', usecols='A:L')

    number1 = df.loc[0]['number 1']
    number2 = df.loc[0]['number 2']
    sol = df.loc[0]['sol']

    context = {'num1': number1,
               'num2': number2,
               'sol': sol}

    return render(request, 'html_pages/excel3.html', context)
