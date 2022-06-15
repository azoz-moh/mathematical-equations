from django.shortcuts import render
from pyexcel_xlsxr import get_data
from reportlab.pdfgen import canvas
from django.http import HttpResponse
import os
import openpyxl
import json
import pyexcel as pe
import formulas
import pandas as pd


# Create your views here.
def index(request):
    df = pd.read_excel(f'{os.getcwd()}/MATH.XLSX', sheet_name='SHEET1', usecols='A:L')
    number1 = df.loc[0]['number 1']
    print(number1)
    number2 = df.loc[0]['number 2']
    print(number2)
    sol = df.loc[0]['sol']
    print(df.loc[0])
    context = {'num1': number1,
               'num2': number2,
               'sol': sol}
    return render(request, 'ttt.html', context)


def ex(request):
    wb = openpyxl.load_workbook(f'{os.getcwd()}\excel\excel2.xlsx')
    sheet1 = wb['Sheet1']
    num1 = request.GET.get('num1')
    num2 = request.GET.get('num2')

    sheet1.cell(column=1, row=2, value=num1)
    sheet1.cell(column=2, row=2, value=num2)
    wb.save(f'{os.getcwd()}/excel/excel2.xlsx')
    wb.close()
    # انشاء نسخة من ملف الاكسيل
    xl_model = formulas.ExcelModel().loads(f'{os.getcwd()}/excel/excel2.XLSX').finish()
    xl_model.calculate()
    # وضع النسخة الملف في المسار الحالي
    xl_model.write(dirpath='./')
    # القراءة من نسخة ملف الاكسيل
    # ملاحظة النسخة تكون بالاحرف الكبيرة
    wb = openpyxl.load_workbook(f'{os.getcwd()}/EXCEL2.XLSX', data_only=True)
    sheet1 = wb['SHEET1']

    df = pd.read_excel(f'{os.getcwd()}/EXCEL2.XLSX', sheet_name='SHEET1', usecols='A:L')
    number1 = df.loc[0]['number 1']
    print(number1)
    number2 = df.loc[0]['number 2']
    print(number2)
    sol = df.loc[0]['sol']
    print(df.loc[0])
    context = {'num1': number1,
               'num2': number2,
               'sol': sol}
    return render(request, 'index_test.html', context)




